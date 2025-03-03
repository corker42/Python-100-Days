import datetime

import openpyxl

# 加载一个工作簿 ---> Workbook
wb = openpyxl.load_workbook('阿里巴巴2020年股票数据.xlsx')
# 获取工作表的名字
print(wb.sheetnames)
# 获取工作表 ---> Worksheet
sheet = wb.worksheets[0]
# 获得单元格的范围
print(sheet.dimensions)
# 获得行数和列数
print(sheet.max_row, sheet.max_column)

# 获取指定单元格的值
print(sheet.cell(3, 3).value)
print(sheet['C3'].value)
print(sheet['G255'].value)

# 获取多个单元格（嵌套元组）
print(sheet['A2:C5'][0][0].value)
print(sheet['A2:C5'])

# # 读取所有单元格的数据
# for row_ch in range(2, sheet.max_row + 1):
#     for col_ch in 'ABCDEFG':
#         value = sheet[f'{col_ch}{row_ch}'].value
#         if type(value) == datetime.datetime:
#             print(value.strftime('%Y年%m月%d日'), end='\t')
#         elif type(value) == int:
#             print(f'{value:<10d}', end='\t')
#         elif type(value) == float:
#             print(f'{value:.4f}', end='\t')
#         else:
#             print(value, end='\t')
#     print()

# import random

# # 第一步：创建工作簿（Workbook）
# wb = openpyxl.Workbook()

# # 第二步：添加工作表（Worksheet）
# sheet = wb.active
# sheet.title = '期末成绩'

# titles = ('姓名', '语文', '数学', '英语')
# for col_index, title in enumerate(titles):
#     sheet.cell(1, col_index + 1, title)

# names = ('关羽', '张飞', '赵云', '马超', '黄忠')
# for row_index, name in enumerate(names):
#     sheet.cell(row_index + 2, 1, name)
#     for col_index in range(2, 5):
#         sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

# # 第四步：保存工作簿
# wb.save('考试成绩表.xlsx')

